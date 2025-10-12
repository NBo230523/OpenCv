import cv2
from mtcnn import MTCNN
from keras_facenet import FaceNet
import joblib
import pickle
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import os
from openpyxl import Workbook, load_workbook
import datetime


SIMILARITY_THRESHOLD = 0.6
CONFIDENCE_THRESHOLD = 0.8

EXCEL_PATH = 'D:/OpenCv/OpenCv/data/diem_danh.xlsx'

def ghi_diem_danh(ten):
    now = datetime.datetime.now()
    time_str = now.strftime("%Y-%m-%d %H:%M:%S")

    # Nếu file chưa có, tạo mới
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "DiemDanh"
        ws.append(["Tên", "Thời gian có mặt"])
        wb.save(EXCEL_PATH)

    # Mở file và ghi dữ liệu
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Kiểm tra xem người này đã được ghi trong 5 phút gần đây chưa (tránh ghi trùng liên tục)
    last_row = None
    for row in ws.iter_rows(values_only=True):
        if row[0] == ten:
            last_row = row
    if last_row:
        last_time = datetime.datetime.strptime(last_row[1], "%Y-%m-%d %H:%M:%S")
        if (now - last_time).total_seconds() < 7200:
            return

    ws.append([ten, time_str])
    wb.save(EXCEL_PATH)

detector = MTCNN()
embedder = FaceNet()

model = joblib.load('D:/OpenCv/OpenCv/data/face_model.pkl')
label_encoder = joblib.load('D:/OpenCv/OpenCv/data/label_encoder.pkl')

with open('D:/OpenCv/OpenCv/data/known_embeddings.pkl', 'rb') as f:
    known = pickle.load(f)

known_embeddings = known['embeddings']
known_names = known['names']

cam = cv2.VideoCapture(0)

print("🎥 Nhấn 'q' để thoát.")
while True:
    ret, frame = cam.read()
    if not ret:
        break

    faces = detector.detect_faces(frame)
    for face in faces:
        x, y, w, h = face['box']
        x, y = max(0, x), max(0, y)
        cropped = frame[y:y+h, x:x+w]
        try:
            resized = cv2.resize(cropped, (160, 160))
            embedding = embedder.embeddings([resized])[0]

            # 1. Dự đoán bằng model
            pred_prob = model.predict_proba([embedding])[0]
            max_index = np.argmax(pred_prob)
            confidence = pred_prob[max_index]

            if confidence > CONFIDENCE_THRESHOLD:
                name = label_encoder.inverse_transform([max_index])[0]
            else:
                # 2. So sánh cosine nếu model không chắc chắn
                sims = cosine_similarity([embedding], known_embeddings)[0]
                best_index = np.argmax(sims)
                best_score = sims[best_index]
                name = known_names[best_index] if best_score > SIMILARITY_THRESHOLD else "Unknown"

            # Ghi điểm danh nếu nhận diện được người hợp lệ
            if name != "Unknown":
                ghi_diem_danh(name)

            # Vẽ khung và tên
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
            cv2.putText(frame, name, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
        except:
            pass

    cv2.imshow("Diem Danh", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
        
cam.release()
cv2.destroyAllWindows()
